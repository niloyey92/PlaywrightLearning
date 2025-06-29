import os

import pytest
from allure_commons.types import AttachmentType
from playwright.sync_api import Page, expect, Browser, HttpCredentials
import  allure
import openpyxl

#from utilities.readingExcel import totalCols


# from utilities.generatingLogs import logger


# def test_user_registration(page):
#     page.goto("https://advantageonlineshopping.com/")
#     #page.set_viewport_size({"width":1920,"height":1080})
#     title = page.title()
#     print(title)
#     assert "Advantage Shopping" in title
#     # logger.info("Title validated")
#     page.wait_for_timeout(2000)
#     page.locator('#hrefUserIcon').click(timeout= 5000)
#     page.get_by_text('CREATE NEW ACCOUNT', exact= True).click()
#     locator = page.locator("//section[@id='registerPage']/article/h3")
#     expect(locator).to_contain_text('CREATE ACCOUNT')
#     page.locator("input[name='usernameRegisterPage']").fill("john150")
#     page.locator("input[name='passwordRegisterPage']").fill("Vanhel1234!")
#     page.locator("input[name='emailRegisterPage']").fill("janhen@example.com")
#     page.locator("input[name='confirm_passwordRegisterPage']").fill("Vanhel1234!")
#     page.locator("input[name='first_nameRegisterPage']").fill("John")
#     page.locator("input[name='last_nameRegisterPage']").fill("Cena")
#     page.locator("input[name='phone_numberRegisterPage']").fill("8768182121")
#     dropdown = page.locator("select[name ='countryListboxRegisterPage']")
#     dropdown.select_option('India')
#     options = page.locator("[name ='countryListboxRegisterPage']>option").all()
#     for option in options:
#         print(option.inner_text())
#     print(len(options))
#     page.locator("input[name='cityRegisterPage']").fill("New Delhi")
#     page.locator("input[name='addressRegisterPage']").fill("123 Jahanen Street, New Delhi")
#     page.locator("input[name='state_/_province_/_regionRegisterPage']").fill("Delhi")
#     page.locator("input[name='postal_codeRegisterPage']").fill("110001")
#     page.locator("input[name ='i_agree']").click()
#     page.locator("#register_btn").click()
#     expect(page.locator("//ul[@class='roboto-light desktop-handler']/li[3]/a/span")).to_contain_text("john150")

# @pytest.mark.usefixtures("log_on_failure")
# @allure.feature("Login test")
# @allure.severity(allure.severity_level.MINOR)
# def test_login_existing_user_success(page):
#     with allure.step("navigate to the url and check the login is success"):
#       page.goto("https://advantageonlineshopping.com/")
#       title = page.title()
#       print(title)
#       allure.attach(page.screenshot(path="screenshot/loginPage1.png"),
#                     name="landed on the homepage",
#                     attachment_type=AttachmentType.PNG)
#       assert "Advantage Shopping" in title
#       page.wait_for_timeout(2000)
#       page.locator('#hrefUserIcon').click(timeout= 5000)
#       page.locator("input[name='username']").fill("john128")
#       page.locator("input[name='password']").fill("Vanhel1234!")
#       allure.attach(page.screenshot(path="screenshot/loginPage2.png"), name="Username and password details are provided",
#                     attachment_type=AttachmentType.PNG)
#       page.get_by_role('button',name="SIGN IN").click()
#       expect(page.locator("//ul[@class='roboto-light desktop-handler']/li[3]/a/span")).to_contain_text("john128")
#     allure.attach(page.screenshot(path="screenshot/loginPage3.png"),name = "Login is success",attachment_type=AttachmentType.PNG)
#     # logger.info("Login is success")

def test_login_random_user_fail(page):
    page.goto("https://advantageonlineshopping.com/")
    title = page.title()
    print(title)
    assert "Advantage Shopping" in title
    page.wait_for_timeout(2000)
    page.locator('#hrefUserIcon').click(timeout=5000)
    page.locator("input[name='username']").fill("john1671")
    page.locator("input[name='password']").fill("Vanhel1234!")
    page.get_by_role('button', name="SIGN IN").click()
    expect(page.locator('#signInResultMessage')).to_contain_text('Incorrect user name or password.')

# def test_authentication_validation_success(page,browser: Browser):
#     context = browser.new_context(
#     http_credentials={"username":"admin","password":"admin"}
#     )
#     page = context.new_page()
#     page.goto("https://the-internet.herokuapp.com/basic_auth")
#     page.wait_for_timeout(5000)
#     expect(page.locator("div[class='example']>h3")).to_contain_text("Basic Auth")

# def test_authentication_validation_fail(page,browser: Browser):
#     context = browser.new_context(
#     http_credentials={"username":"admin1","password":"admin"}
#     )
#     page = context.new_page()
#     page.goto("https://the-internet.herokuapp.com/basic_auth")
#     page.wait_for_timeout(5000)
#     expect(page.locator("//body[contains (text(),'Not authorized')]")).to_be_visible()

def get_data():
  workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
  sheet = workbook["LoginTest"]
  totalRows = sheet.max_row
  totalCols = sheet.max_column
  mainList=[]
  print(totalCols)
  print(totalRows)
  for i in range (2,totalRows+1):
      dataList=[]
      for j in range(1, totalCols + 1):
          data = sheet.cell(row=i,column=j).value
          dataList.insert(j,data)
      mainList.insert(i,dataList)
  print(mainList)
  return mainList


@allure.feature("Login test")
@allure.severity(allure.severity_level.MINOR)
@pytest.mark.parametrize("username,password",get_data())
def test_login_existing_multiple_user_success(page,username,password):
    with allure.step("navigate to the url and check the login is success"):
      page.goto("https://advantageonlineshopping.com/")
      title = page.title()
      print(title)
      allure.attach(page.screenshot(path="screenshot/loginPage1.png"),
                    name="landed on the homepage",
                    attachment_type=AttachmentType.PNG)
      assert "Advantage Shopping" in title
      page.wait_for_timeout(2000)
      page.locator('#hrefUserIcon').click(timeout= 5000)
      page.locator("input[name='username']").fill(username)
      page.locator("input[name='password']").fill(password)
      allure.attach(page.screenshot(path="screenshot/loginPage2.png"), name="Username and password details are provided",
                    attachment_type=AttachmentType.PNG)
      page.get_by_role('button',name="SIGN IN").click()
      expect(page.locator("//ul[@class='roboto-light desktop-handler']/li[3]/a/span")).to_contain_text(username)
    allure.attach(page.screenshot(path="screenshot/loginPage3.png"),name = "Login is success",attachment_type=AttachmentType.PNG)

# def test_file_upload(page):
#     page.goto('https://the-internet.herokuapp.com/upload');
#     page.locator('#file-upload').set_input_files("C:\\Users\\JS492UT\\Downloads\\myFile.pdf")
#     page.locator('#file-submit').click(timeout= 12000);
#
# def test_multiple_file_upload(page):
#     page.goto('https://www.w3schools.com/jsref/tryit.asp?filename=tryjsref_fileupload_multiple',timeout=40000);
#     frame = page.frame_locator("iframeResult")
#     frame.locator("#myFile").set_input_files(["C:\\Users\\JS492UT\\Downloads\\myFile.pdf","C:\\Users\\JS492UT\\Downloads\\myFile2.pdf"])
#
# def test_download_file(page):
#     page.goto('https://the-internet.herokuapp.com/download',timeout=10000);
#     with page.expect_download() as download_file:
#         page.locator('#content > div > a:nth-child(2)').click()
#         download = download_file.value
#         project_directory = os.path.join(os.path.dirname(os.getcwd()),"downloads")
#         os.makedirs(project_directory,exist_ok=True)
#         file_path = os.path.join(project_directory,"Bus_Ticket_TTB714817104.pdf")
#         download.save_as(file_path)
#         print(f"File downloaded : {file_path}")
#
# def test_web_tables(page):
#     page.goto("https://cosmocode.io/automation-practice-webtable/",timeout=60000)
#
#     row_count = page.locator("#countries > tbody > tr").count()
#     print("Row count is ",row_count)
#     col_count = page.locator("#countries > tbody > tr:nth-child(2) > td").count()
#     print("Column count is ", col_count)
#     all_inner_text = page.locator("#countries > tbody > tr").all_inner_texts()
#     for inner_text in all_inner_text:
#         print(inner_text)
#
# def test_window_handles(page):
#     page.goto("https://the-internet.herokuapp.com/windows")
#     print(page.title())
#     expect(page).to_have_title("The Internet")
#     page.locator(".example>a").click(timeout=60000)
#     with page.expect_popup() as popup_info:
#         new_page = popup_info.value
#         new_page.wait_for_load_state()
#         print(new_page.title())
#         expect(new_page).to_have_title("New Window",timeout=60000)
