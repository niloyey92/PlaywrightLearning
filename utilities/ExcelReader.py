import openpyxl

def getRowCount(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(path, sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value


def setCellData(path, sheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)

path = "..//excel//testdata.xlsx"
sheetName = "LoginTest"

rows= getRowCount(path,sheetName)
print (rows)
print (getColCount(path,sheetName))

print(getCellData(path,sheetName,3,2))