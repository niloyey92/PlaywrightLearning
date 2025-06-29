import openpyxl

workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
sheet = workbook["LoginTest"]
# sheet.cell(row=1,column=3).value="age"


for rows in range(6,10):
    for cols in range(1,4):
        sheet.cell(row=rows, column=cols).value="Hello"

workbook.save("..//excel//testdata.xlsx")