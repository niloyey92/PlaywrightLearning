import openpyxl

workbook = openpyxl.load_workbook("..//excel//testdata.xlsx")
sheet = workbook["LoginTest"]
totalRows = sheet.max_row
totalCols = sheet.max_column
# print("total rows are ",str(totalRows)," and total cols are ",str(totalCols))
# print(sheet.cell(row=3,column=1).value)
mainList = []
for rows in range(2,totalRows+1):
    dataList = []
    for cols in range(1,totalCols+1):
        # print(sheet.cell(row=rows, column=cols).value,end="    ")
        data = sheet.cell(row=rows, column=cols).value
        print(data)
        dataList.insert(cols-1, data)
    mainList.insert(rows-2, dataList)
print(mainList)